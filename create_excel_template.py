from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

output = '/home/ubuntu/MARSB-GYM-modelo-novo-treino.xlsx'
wb = Workbook()
ws = wb.active
ws.title = 'Treinos'

headers = ['Treino', 'Nome do treino', 'Foco', 'Exercício', 'Séries', 'Repetições', 'Descanso (seg)', 'Técnica / nota']
rows = [
    ['A', 'Peito + Core', 'Peitoral e fortalecimento abdominal', 'Supino inclinado com halteres', 4, '6 a 8', 150, 'Controle na fase excêntrica'],
    ['A', 'Peito + Core', 'Peitoral e fortalecimento abdominal', 'Crossover na polia', 4, '10 a 12', 90, 'Contração máxima'],
    ['B', 'Costas + Core', 'Dorsais e estabilização', 'Puxada frontal', 4, '8 a 10', 120, 'Peito aberto'],
    ['B', 'Costas + Core', 'Dorsais e estabilização', 'Remada baixa', 4, '8 a 12', 120, 'Sem impulso'],
    ['C', 'Pernas', 'Quadríceps, posteriores e glúteos', 'Agachamento livre', 4, '6 a 8', 180, 'Amplitude confortável'],
    ['C', 'Pernas', 'Quadríceps, posteriores e glúteos', 'Leg press', 4, '10 a 12', 120, 'Não perder a lombar'],
    ['D', 'Ombros + Braços', 'Deltoides, bíceps e tríceps', 'Desenvolvimento com halteres', 4, '8 a 10', 120, 'Movimento controlado'],
    ['D', 'Ombros + Braços', 'Deltoides, bíceps e tríceps', 'Rosca direta', 3, '10 a 12', 90, 'Cotovelos estáveis'],
]

header_fill = PatternFill('solid', fgColor='007AFF')
header_font = Font(color='FFFFFF', bold=True)
thin = Side(style='thin', color='D1D1D6')
for col, value in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=value)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=thin)

for r, row in enumerate(rows, 2):
    for c, value in enumerate(row, 1):
        cell = ws.cell(row=r, column=c, value=value)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = Border(bottom=thin)

widths = [10, 28, 42, 40, 10, 14, 16, 34]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:H{len(rows) + 1}'
tab = Table(displayName='TabelaTreinos', ref=f'A1:H{len(rows) + 1}')
tab.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

validation = DataValidation(type='list', formula1='"A,B,C,D"', allow_blank=False)
validation.error = 'Use somente A, B, C ou D.'
validation.errorTitle = 'Treino inválido'
ws.add_data_validation(validation)
validation.add(f'A2:A500')

positive = DataValidation(type='whole', operator='between', formula1='1', formula2='10', allow_blank=False)
positive.error = 'Informe um número inteiro entre 1 e 10.'
ws.add_data_validation(positive)
positive.add('E2:E500')

rest = DataValidation(type='whole', operator='between', formula1='0', formula2='600', allow_blank=False)
rest.error = 'Informe um descanso entre 0 e 600 segundos.'
ws.add_data_validation(rest)
rest.add('G2:G500')

info = wb.create_sheet('Instruções')
info_rows = [
    ['Como preencher a planilha do MARSB-GYM'],
    ['Cada linha da aba Treinos representa um exercício.'],
    ['Mantenha a primeira linha com os títulos das colunas.'],
    ['Treino: use somente A, B, C ou D.'],
    ['Séries: número inteiro entre 1 e 10.'],
    ['Repetições: use um número ou uma faixa, por exemplo 8 ou 8 a 12.'],
    ['Descanso (seg): informe o intervalo entre 0 e 600 segundos.'],
    ['Depois de preencher, abra o MARSB-GYM, entre em Configuração > Plano e toque em Importar planilha Excel.'],
    ['A importação substitui o plano personalizado atual; exporte um backup antes se quiser preservá-lo.'],
]
for r, row in enumerate(info_rows, 1):
    cell = info.cell(row=r, column=1, value=row[0])
    cell.alignment = Alignment(wrap_text=True, vertical='top')
    if r == 1:
        cell.fill = header_fill
        cell.font = header_font
    info.row_dimensions[r].height = 28 if r > 1 else 32
info.column_dimensions['A'].width = 110
info.freeze_panes = 'A2'

wb.properties.title = 'Modelo de novo treino - MARSB-GYM'
wb.properties.subject = 'Importação de treinos por planilha'
wb.save(output)
print(output)
print(f'{len(rows)} linhas de exercícios + cabeçalho')
